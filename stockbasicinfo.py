# -*- coding: utf-8-*-
# stock basic information from www.eastmoney.com
# By Jun Lv, 2016, Jun.Alex.Lv@gmail.com
# This software is not placed into the public domain
# Revision date: Aug 11, 2016
#Python version:2.7.8
# Version: 1.0

#import psycopg2
import urllib2
import multiprocessing
import sys
import time
import re
import os
import logging
import argparse
import warnings
import csv
from  openpyxl.reader.excel  import  load_workbook 



# 从数据库根据股票列表生成二维数组供其它模块调用
def StockList(excelfile):
    stockid=[]
    head=''
    try:
        warnings.simplefilter("ignore")
        wb = load_workbook(excelfile)  
        ws = wb.get_sheet_by_name("stockbasicinfo") 
        warnings.simplefilter("default")
        lastrow=ws.max_row
        for rowno in range(2,lastrow+1):
            if(ws.cell(row=rowno,column=1).value[0]=='0' or ws.cell(row=rowno,column=1).value[0]=='2' or ws.cell(row=rowno,column=1).value[0]=='3'):
                head='sz'
            if(ws.cell(row=rowno,column=1).value[0]=='6' or ws.cell(row=rowno,column=1).value[0]=='9' ):
                head='sh'
            stockid.append(head+ws.cell(row=rowno,column=1).value)
        
        return stockid

    except:
        exceptionType, exceptionValue, exceptionTraceback = sys.exc_info()
        sys.exit("file handle failed!\n ->%s" % (exceptionValue))   

def DownLoad(stkInfo,i,option,fileNum,processNum,p):
    #远程debug使用   
    import wingdbstub
    time.sleep(10)
    try:
        wingdbstub.Ensure()
        print 'Connected to wingIDE, No %d#, PID#%s' % (i,os.getpid())
    except ValueError:
        error=1   
    
    if (p !=''):
        proxy = urllib2.ProxyHandler({'http': '%s' % p})
        opener = urllib2.build_opener(proxy)
        urllib2.install_opener(opener)        
    
    #reportType={'1':'BalanceSheet','2':'ProfitStatement','3':'CashFlow'}
    #计算单进程所需要处理的文件个数,考虑N个进程，将文件数fileNum除以进程数ProcessNum，余值modNum再平均分配到前modNum个进程。
    #即：前modNum个进程处理文件数,start=i*(avgNum+1),end=(i+1)*(avgNum+1)，大于等于modNum进程处理文件数，start=i*avgNum+modNum,end=(i+1)*avgNum+modNum
    #filenum=num/processNum    
    
    avgNum=fileNum/processNum                          #进程平均下载文件数，可能有余值
    modNum=fileNum%processNum                       #余值文件数
    
    if(i<modNum):
        start=i*(avgNum+1)
        end=(i+1)*(avgNum+1)
    else:
        start=i*avgNum+modNum
        end=(i+1)*avgNum+modNum
    
    starttime = time.time()
    index=start
    failcount=0

    while index<end:
        try:
            #url='http://vip.stock.finance.sina.com.cn/corp/go.php/vCI_CorpInfo/stockid/%s.phtml'  % stkInfo[index]
            url='http://f10.eastmoney.com/f10_v2/CompanySurvey.aspx?code=%s' % stkInfo[index]
            filename=('%s_basicinfo.html' % (stkInfo[index])).decode('utf-8').encode('cp936')
            content=urllib2.urlopen(url,timeout=10)
            with open('e:/My Stock/Data/stockbasicinfo/%s'  % (filename), mode='w') as targetfile:
                targetfile.write(content.read())
                index=index+1
                failcount=0
                       
        except:
            exceptionType, exceptionValue, exceptionTraceback = sys.exc_info()
            #sys.exit("file %s download failed! belong to #%d process%s->exceptionTraceback: %s,exceptionValue:%s" % (filename,i,os.getpid(),exceptionTraceback,exceptionValue))      
            failcount +=1
            print ("index=%d,failcount=%d" % (index,failcount))    
            
    endtime = time.time()
    print ("%d#:PID# %s\tElapsed Time : %d" %(i,os.getpid(),endtime-starttime))    
    
    
def Save2xlsx(excelfile):
    from bs4 import BeautifulSoup
    stockid=''
    head=''
    try:
        warnings.simplefilter("ignore")
        wb = load_workbook(excelfile)  
        ws = wb.get_sheet_by_name("stockbasicinfo") 
        warnings.simplefilter("default")             
        
        lastrow=ws.max_row
        #lastrow=2
        for rowno in range(2,lastrow+1):
            if(ws.cell(row=rowno,column=1).value[0]=='0' or ws.cell(row=rowno,column=1).value[0]=='2' or ws.cell(row=rowno,column=1).value[0]=='3'):
                head='sz'
            if(ws.cell(row=rowno,column=1).value[0]=='6' or ws.cell(row=rowno,column=1).value[0]=='9' ):
                head='sh'
            stockid=(head+ws.cell(row=rowno,column=1).value)
            with open('e:/My Stock/Data/stockbasicinfo/%s_basicinfo.html'  % (stockid), mode='rb') as f:
                soup=BeautifulSoup(f,"html.parser")
                table = soup.findAll("table")
                #公司名称(CompanyName)
                rows = table[0].findAll('tr')[0]
                cols=rows.findAll("td")[0]
                ws.cell(row=rowno,column=9).value=cols.text
                #公司英文名称(CompanyNameEng)
                rows = table[0].findAll('tr')[1]
                cols=rows.findAll("td")[0]
                ws.cell(row=rowno,column=10).value=cols.text

                #公司网址(webpage)
                rows = table[0].findAll('tr')[11]
                cols=rows.findAll("td")[1]
                ws.cell(row=rowno,column=8).value=cols.text  
                #地区(Area)
                rows = table[0].findAll('tr')[13]
                cols=rows.findAll("td")[0]                
                ws.cell(row=rowno,column=4).value=cols.text       
                
                #上市日期(IPOdate)
                rows = table[1].findAll('tr')[0]
                cols=rows.findAll("td")[1]
                ws.cell(row=rowno,column=5).value=cols.text   
                print 'now the securities code is %s' % stockid,
                sys.stdout.write("\r")
        wb.save(excelfile)
 

    except:
        exceptionType, exceptionValue, exceptionTraceback = sys.exc_info()
        sys.exit("file handle failed!\n ->%s" % (exceptionValue))   
    
    
def Import2DB():
    print 'hello'
    
    
    
def Excel2csv(filename):   
    try:
        warnings.simplefilter("ignore")
        wb = load_workbook(filename)  
        ws = wb.get_sheet_by_name("stockbasicinfo") 
        warnings.simplefilter("default")        
        csv_filename=filename.replace('xlsx','csv')     
        csv_file = file(csv_filename, 'wb')
        csv_file_writer = csv.writer(csv_file)

        for row in ws.rows:
            row_container = []
            for cell in row:
                if type(cell.value) == unicode:
                    row_container.append(cell.value.encode('utf-8'))
                else:
                    row_container.append(str(cell.value))
            csv_file_writer.writerow(row_container)
        csv_file.close()

    except Exception as e:
        print(e)

if __name__== '__main__':   
    # 参数处理说明
    parser = argparse.ArgumentParser(description='Process stock basic infomation data')
    parser.add_argument('--version', action='version', version='%(prog)s 0.1')
    parser.add_argument("-i", metavar="inputFile",default='stockbasicinfo.xlsx',help="stockbasicinfo file name")  
    parser.add_argument('-p',nargs='?',metavar='Proxy add.',const='10.144.1.10:8080',
                        help='using proxy (default proxy: 10.144.1.10:8080)')    
    
    args = parser.parse_args()
    proxy=''
    
    if (args.p != None):
            pattern=('((2[0-4]\d|25[0-5]|[01]?\d\d?)\.){3}(2[0-4]\d|25[0-5]|[01]?\d\d?):(([1-9]\d{0,3})|([1-5]\d{4})|(6[0-4]\d{3})|(65[0-4]\d{2})|(655[0-2]\d)|(6553[0-5]))$')
            r = re.match(pattern,args.p)
            proxy=args.p
             
            if r == None:
                    sys.exit("Wrong proxy address.")                
                    
    if(args.i!=None):
        working_dir="e:\\My Stock\\Data\\"
        infile=os.path.join(working_dir,args.i)        
        if (os.path.exists(infile))==False:
            print "No such directory or file exsits"    
            

      
    # 功能选择菜单
    function={'1':'下载信息','2':'导入数据库'}
    print('Function List:')
    for key in sorted(function.keys()):
            print('{:>4}.\t{:<10}'.format(key, function[key]).decode('utf-8').encode('cp936'))              #格式化字符串
     
    r = None                                                                                                                            #检查输入是否合规
    while r == None:
            option=raw_input('Please choose function:(1-2)')
            r = re.match("^([1-2]).*",option)    
    
    if option=='1':
        #从数据库调取股票列表信息及股票总数
        stkInfo=StockList(infile)
        fileNum=len(stkInfo)
        processNum=input('Enter the process num:')   
        logger=multiprocessing.log_to_stderr()    
        logger.setLevel(logging.INFO)            
        print("Total %d files need to be downloaded!" %fileNum)
        print 'Now, Downloading is begining'
    
        for i in range(processNum):
            p = multiprocessing.Process(target=DownLoad, args=(stkInfo,i,option,fileNum,processNum,proxy))  
            p.start()
    
    if option=='2':
        #Save2xlsx(infile)
        Excel2csv(infile)
   
   
    

